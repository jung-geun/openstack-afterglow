"""GPU 장치 카탈로그 API 단위 테스트 (admin_gpu.py /gpu-devices 엔드포인트).

검증 항목:
1. 모든 엔드포인트: non-admin → 403
2. 변경 계열 엔드포인트: DB 미초기화 → 503
3. 입력 검증: vendor_id/device_id/alias 형식 거부
4. GET 병합 카탈로그: builtin source 표시
5. DELETE: builtin 항목 삭제 거부(409), 없는 항목 404
6. CSV parse_csv: 헤더 스킵, 세미콜론 alias, 행 번호 포함 오류
7. apply_db_overlay: PCI_DEVICE_MAP in-place 반영 + 캐시 무효화
"""

from unittest.mock import AsyncMock, patch

import pytest

from app.services.gpu_catalog import parse_csv, validate_entry
from app.services.gpu_inventory import PCI_DEVICE_MAP, apply_db_overlay

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 인증 (admin-only)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.mark.asyncio
async def test_list_gpu_devices_requires_admin(non_admin_client):
    resp = await non_admin_client.get("/api/admin/gpu-devices")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_upsert_gpu_device_requires_admin(non_admin_client):
    resp = await non_admin_client.post(
        "/api/admin/gpu-devices",
        json={"vendor_id": "10DE", "device_id": "FFFF", "name": "Test GPU", "aliases": ["TESTGPU"]},
    )
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_delete_gpu_device_requires_admin(non_admin_client):
    resp = await non_admin_client.delete("/api/admin/gpu-devices/10DE/FFFF")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_import_gpu_devices_requires_admin(non_admin_client):
    resp = await non_admin_client.post(
        "/api/admin/gpu-devices/import",
        files={"file": ("catalog.csv", b"10DE,FFFF,Test,false,TESTGPU", "text/csv")},
    )
    assert resp.status_code == 403


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DB 미초기화 → 503
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.mark.asyncio
async def test_upsert_gpu_device_db_not_initialized(admin_client):
    with patch("app.database.is_db_available", return_value=False):
        resp = await admin_client.post(
            "/api/admin/gpu-devices",
            json={"vendor_id": "10DE", "device_id": "FFFF", "name": "Test GPU", "aliases": ["TESTGPU"]},
        )
    assert resp.status_code == 503


@pytest.mark.asyncio
async def test_delete_gpu_device_db_not_initialized(admin_client):
    with patch("app.database.is_db_available", return_value=False):
        resp = await admin_client.delete("/api/admin/gpu-devices/10DE/FFFF")
    assert resp.status_code == 503


@pytest.mark.asyncio
async def test_import_gpu_devices_db_not_initialized(admin_client):
    with patch("app.database.is_db_available", return_value=False):
        resp = await admin_client.post(
            "/api/admin/gpu-devices/import",
            files={"file": ("catalog.csv", b"10DE,FFFF,Test,false,TESTGPU", "text/csv")},
        )
    assert resp.status_code == 503


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GET 병합 카탈로그 (DB 없이도 builtin/config 표시)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.mark.asyncio
async def test_list_gpu_devices_returns_builtin(admin_client):
    with patch("app.database.is_db_available", return_value=False):
        resp = await admin_client.get("/api/admin/gpu-devices")
    assert resp.status_code == 200
    devices = resp.json()["devices"]
    assert devices, "내장 기본값 카탈로그가 비어있으면 안 됨"
    rtx3090 = next(d for d in devices if d["vendor_id"] == "10DE" and d["device_id"] == "2204")
    assert rtx3090["name"] == "RTX 3090"
    assert rtx3090["source"] == "builtin"
    assert rtx3090["vendor_name"] == "NVIDIA"
    assert "RTX3090" in rtx3090["aliases"]


@pytest.mark.asyncio
async def test_list_gpu_devices_marks_db_source(admin_client):
    db_entries = [
        {"vendor_id": "10DE", "device_id": "2204", "name": "RTX 3090 custom", "is_audio": False, "aliases": []}
    ]
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.list_db_devices", new=AsyncMock(return_value=db_entries)),
    ):
        resp = await admin_client.get("/api/admin/gpu-devices")
    assert resp.status_code == 200
    rtx3090 = next(d for d in resp.json()["devices"] if d["device_id"] == "2204")
    assert rtx3090["source"] == "db"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# POST 단건 upsert + 입력 검증
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.mark.asyncio
async def test_upsert_gpu_device_ok(admin_client):
    saved = {"vendor_id": "10DE", "device_id": "FFFF", "name": "Test GPU", "is_audio": False, "aliases": ["TESTGPU"]}
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.upsert_device", new=AsyncMock(return_value=saved)) as mock_upsert,
        patch("app.services.gpu_catalog.refresh_device_map_from_db", new=AsyncMock()) as mock_refresh,
    ):
        resp = await admin_client.post(
            "/api/admin/gpu-devices",
            json={"vendor_id": "10DE", "device_id": "FFFF", "name": "Test GPU", "aliases": ["TESTGPU"]},
        )
    assert resp.status_code == 200
    assert resp.json()["device_id"] == "FFFF"
    mock_upsert.assert_awaited_once()
    mock_refresh.assert_awaited_once()


@pytest.mark.asyncio
async def test_upsert_gpu_device_rejects_bad_vendor_id(admin_client):
    with patch("app.database.is_db_available", return_value=True):
        resp = await admin_client.post(
            "/api/admin/gpu-devices",
            json={"vendor_id": "XYZ!", "device_id": "FFFF", "name": "Test", "aliases": []},
        )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_upsert_gpu_device_rejects_bad_alias(admin_client):
    with patch("app.database.is_db_available", return_value=True):
        resp = await admin_client.post(
            "/api/admin/gpu-devices",
            json={"vendor_id": "10DE", "device_id": "FFFF", "name": "Test", "aliases": ["bad;alias\nrm -rf"]},
        )
    assert resp.status_code == 400


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DELETE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.mark.asyncio
async def test_delete_gpu_device_ok(admin_client):
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.delete_device", new=AsyncMock(return_value=True)),
        patch("app.services.gpu_catalog.refresh_device_map_from_db", new=AsyncMock()) as mock_refresh,
    ):
        resp = await admin_client.delete("/api/admin/gpu-devices/10DE/FFFF")
    assert resp.status_code == 204
    mock_refresh.assert_awaited_once()


@pytest.mark.asyncio
async def test_delete_gpu_device_builtin_conflict(admin_client):
    """builtin 항목(RTX 3090)은 DB에 없으므로 삭제 시 409."""
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.delete_device", new=AsyncMock(return_value=False)),
    ):
        resp = await admin_client.delete("/api/admin/gpu-devices/10DE/2204")
    assert resp.status_code == 409


@pytest.mark.asyncio
async def test_delete_gpu_device_not_found(admin_client):
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.delete_device", new=AsyncMock(return_value=False)),
    ):
        resp = await admin_client.delete("/api/admin/gpu-devices/10DE/FFFF")
    assert resp.status_code == 404


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 템플릿 export (다운로드)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.mark.asyncio
async def test_export_gpu_devices_requires_admin(non_admin_client):
    resp = await non_admin_client.get("/api/admin/gpu-devices/export")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_gpu_devices_csv(admin_client):
    with patch("app.database.is_db_available", return_value=False):
        resp = await admin_client.get("/api/admin/gpu-devices/export?format=csv")
    assert resp.status_code == 200
    assert 'filename="gpu_devices.csv"' in resp.headers["content-disposition"]
    text = resp.content.decode("utf-8-sig")
    lines = text.splitlines()
    assert lines[0] == "vendor_id,device_id,name,is_audio,aliases,source"
    assert any(line.startswith("10DE,2204,RTX 3090,false,") for line in lines)


@pytest.mark.asyncio
async def test_export_gpu_devices_xlsx_roundtrip(admin_client):
    """export한 xlsx를 그대로 import 파서에 넣으면 동일 항목이 복원된다 (source 컬럼 무시)."""
    from openpyxl import load_workbook

    from app.services.gpu_catalog import parse_xlsx

    with patch("app.database.is_db_available", return_value=False):
        resp = await admin_client.get("/api/admin/gpu-devices/export")
    assert resp.status_code == 200
    assert 'filename="gpu_devices.xlsx"' in resp.headers["content-disposition"]

    import io

    wb = load_workbook(io.BytesIO(resp.content))
    header = [c.value for c in next(wb.worksheets[0].iter_rows(max_row=1))]
    assert header == ["vendor_id", "device_id", "name", "is_audio", "aliases", "source"]

    entries = parse_xlsx(resp.content)
    rtx3090 = next(e for e in entries if e["device_id"] == "2204")
    assert rtx3090["name"] == "RTX 3090"
    assert "RTX3090" in rtx3090["aliases"]
    assert rtx3090["is_audio"] is False


@pytest.mark.asyncio
async def test_export_gpu_devices_invalid_format(admin_client):
    resp = await admin_client.get("/api/admin/gpu-devices/export?format=pdf")
    assert resp.status_code == 422


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CSV import 엔드포인트
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_CSV_OK = (
    b"vendor_id,device_id,name,is_audio,aliases\n10DE,AAAA,Test GPU A,false,TESTA;testa\n10DE,BBBB,Test GPU B,true,\n"
)


@pytest.mark.asyncio
async def test_import_gpu_devices_replace_default(admin_client):
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.bulk_import", new=AsyncMock(return_value=2)) as mock_bulk,
        patch("app.services.gpu_catalog.refresh_device_map_from_db", new=AsyncMock()),
    ):
        resp = await admin_client.post(
            "/api/admin/gpu-devices/import",
            files={"file": ("catalog.csv", _CSV_OK, "text/csv")},
        )
    assert resp.status_code == 200
    assert resp.json() == {"imported": 2, "mode": "replace"}
    assert mock_bulk.await_args.kwargs.get("mode") == "replace"


@pytest.mark.asyncio
async def test_import_gpu_devices_upsert_mode(admin_client):
    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.bulk_import", new=AsyncMock(return_value=2)) as mock_bulk,
        patch("app.services.gpu_catalog.refresh_device_map_from_db", new=AsyncMock()),
    ):
        resp = await admin_client.post(
            "/api/admin/gpu-devices/import?mode=upsert",
            files={"file": ("catalog.csv", _CSV_OK, "text/csv")},
        )
    assert resp.status_code == 200
    assert resp.json()["mode"] == "upsert"
    assert mock_bulk.await_args.kwargs.get("mode") == "upsert"


@pytest.mark.asyncio
async def test_import_gpu_devices_invalid_mode(admin_client):
    with patch("app.database.is_db_available", return_value=True):
        resp = await admin_client.post(
            "/api/admin/gpu-devices/import?mode=overwrite",
            files={"file": ("catalog.csv", _CSV_OK, "text/csv")},
        )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_import_gpu_devices_xlsx(admin_client):
    """엑셀(xlsx) 업로드도 CSV와 동일하게 동작한다."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["vendor_id", "device_id", "name", "is_audio", "aliases", "source"])
    ws.append(["10DE", "AAAA", "Test GPU A", "false", "TESTA;testa", "db"])
    ws.append(["10DE", "BBBB", "Test GPU B", True, "", ""])  # bool 셀도 허용
    buf = io.BytesIO()
    wb.save(buf)

    captured = {}

    async def _fake_bulk(entries, mode):
        captured["entries"] = entries
        captured["mode"] = mode
        return len(entries)

    with (
        patch("app.database.is_db_available", return_value=True),
        patch("app.services.gpu_catalog.bulk_import", new=AsyncMock(side_effect=_fake_bulk)),
        patch("app.services.gpu_catalog.refresh_device_map_from_db", new=AsyncMock()),
    ):
        resp = await admin_client.post(
            "/api/admin/gpu-devices/import",
            files={
                "file": (
                    "gpu_devices.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert resp.status_code == 200
    assert resp.json() == {"imported": 2, "mode": "replace"}
    assert captured["entries"][0]["aliases"] == ["TESTA", "testa"]
    assert captured["entries"][1]["is_audio"] is True


@pytest.mark.asyncio
async def test_import_gpu_devices_bad_csv_returns_row_number(admin_client):
    bad = b"10DE,ZZZZ,Bad Device,false,ALIAS\n"
    with patch("app.database.is_db_available", return_value=True):
        resp = await admin_client.post(
            "/api/admin/gpu-devices/import",
            files={"file": ("catalog.csv", bad, "text/csv")},
        )
    assert resp.status_code == 400
    assert "1행" in resp.json()["detail"]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 서비스 단위: validate_entry / parse_csv
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def test_validate_entry_normalizes_hex_case():
    assert validate_entry("10de", "2b85", "RTX 5090", ["RTX5090"]) == ("10DE", "2B85")


@pytest.mark.parametrize(
    "vendor,device,name,aliases",
    [
        ("10D", "2204", "x", []),  # vendor 3자리
        ("10DE", "22045", "x", []),  # device 5자리
        ("10DE", "2204", "", []),  # 빈 name
        ("10DE", "2204", "x", ["evil$(rm)"]),  # alias 특수문자
        ("10DE", "2204", "x", ["a" * 65]),  # alias 길이 초과
    ],
)
def test_validate_entry_rejects(vendor, device, name, aliases):
    with pytest.raises(ValueError):
        validate_entry(vendor, device, name, aliases)


def test_parse_csv_ok():
    entries = parse_csv("vendor_id,device_id,name,is_audio,aliases\n10de,AAAA,Test GPU,false,TESTA;testa\n")
    assert entries == [
        {"vendor_id": "10de", "device_id": "AAAA", "name": "Test GPU", "is_audio": False, "aliases": ["TESTA", "testa"]}
    ]


def test_parse_csv_skips_blank_lines_and_header():
    entries = parse_csv("vendor_id,device_id,name,is_audio,aliases\n\n10DE,AAAA,GPU A,true,\n")
    assert len(entries) == 1
    assert entries[0]["is_audio"] is True
    assert entries[0]["aliases"] == []


def test_parse_csv_error_includes_row_number():
    with pytest.raises(ValueError, match="3행"):
        parse_csv("vendor_id,device_id,name,is_audio,aliases\n10DE,AAAA,GPU A,false,\n10DE,XX,Bad,false,\n")


def test_parse_csv_empty_raises():
    with pytest.raises(ValueError):
        parse_csv("vendor_id,device_id,name,is_audio,aliases\n")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# apply_db_overlay — PCI_DEVICE_MAP in-place 갱신
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.fixture
def _restore_device_map():
    yield
    apply_db_overlay([])  # base map(내장 + config)으로 복원


def test_apply_db_overlay_adds_and_reverts(_restore_device_map):
    assert "EEEE" not in PCI_DEVICE_MAP.get("10DE", {})
    apply_db_overlay(
        [{"vendor_id": "10de", "device_id": "eeee", "name": "Custom GPU", "is_audio": False, "aliases": ["CUSTOM"]}]
    )
    assert PCI_DEVICE_MAP["10DE"]["EEEE"]["name"] == "Custom GPU"
    # 내장 항목은 유지
    assert PCI_DEVICE_MAP["10DE"]["2204"]["name"] == "RTX 3090"
    apply_db_overlay([])
    assert "EEEE" not in PCI_DEVICE_MAP["10DE"]


def test_apply_db_overlay_overrides_builtin(_restore_device_map):
    apply_db_overlay(
        [
            {
                "vendor_id": "10DE",
                "device_id": "2204",
                "name": "RTX 3090 Renamed",
                "is_audio": False,
                "aliases": ["R3090"],
            }
        ]
    )
    assert PCI_DEVICE_MAP["10DE"]["2204"]["name"] == "RTX 3090 Renamed"
    apply_db_overlay([])
    assert PCI_DEVICE_MAP["10DE"]["2204"]["name"] == "RTX 3090"


def test_invalidate_norm_map_picks_up_overlay(_restore_device_map):
    from app.services.gpu_quota import invalidate_norm_map, normalize_gpu_alias

    apply_db_overlay(
        [
            {
                "vendor_id": "10DE",
                "device_id": "EEEE",
                "name": "Custom GPU",
                "is_audio": False,
                "aliases": ["MYGPU", "my-gpu"],
            }
        ]
    )
    invalidate_norm_map()
    assert normalize_gpu_alias("my-gpu") == "MYGPU"
    apply_db_overlay([])
    invalidate_norm_map()
    assert normalize_gpu_alias("my-gpu") == "my-gpu"
