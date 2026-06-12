"""GPU PCI 장치 카탈로그 DB CRUD + CSV 일괄 import.

DB 항목은 내장 기본값(_DEFAULT_PCI_DEVICE_MAP) + config.toml [[gpu.devices]] 위에
overlay되어 PCI_DEVICE_MAP에 반영된다 (gpu_inventory.apply_db_overlay).
DB가 초기화되어 있어야 동작 (is_db_available() 확인 후 호출).
"""

from __future__ import annotations

import csv
import io
import logging
import re

from sqlalchemy import delete, select

from app.database import get_session_factory
from app.models.db import GpuDeviceCatalog

_logger = logging.getLogger(__name__)

_HEX4_RE = re.compile(r"^[0-9A-Fa-f]{4}$")
_ALIAS_RE = re.compile(r"^[A-Za-z0-9 _\-]{1,64}$")

CSV_COLUMNS = ["vendor_id", "device_id", "name", "is_audio", "aliases"]


def validate_entry(vendor_id: str, device_id: str, name: str, aliases: list[str]) -> tuple[str, str]:
    """카탈로그 항목 필드 검증. 정규화된 (vendor_id, device_id)를 반환.

    alias는 flavor extra spec(pci_passthrough:alias)과 nova 설정으로 흘러가는
    값이므로 화이트리스트 정규식으로 검증한다.
    """
    if not _HEX4_RE.match(vendor_id or ""):
        raise ValueError(f"vendor_id는 4자리 16진수여야 합니다: {vendor_id!r}")
    if not _HEX4_RE.match(device_id or ""):
        raise ValueError(f"device_id는 4자리 16진수여야 합니다: {device_id!r}")
    if not name or len(name) > 128:
        raise ValueError("name은 1~128자여야 합니다")
    for alias in aliases:
        if not _ALIAS_RE.match(alias):
            raise ValueError(f"alias 형식이 올바르지 않습니다 (영숫자/공백/_/- 1~64자): {alias!r}")
    return vendor_id.upper(), device_id.upper()


def _row_to_dict(row: GpuDeviceCatalog) -> dict:
    return {
        "vendor_id": row.vendor_id,
        "device_id": row.device_id,
        "name": row.name,
        "is_audio": row.is_audio,
        "aliases": list(row.aliases or []),
    }


async def list_db_devices() -> list[dict]:
    """DB 카탈로그 항목 전체 조회."""
    factory = get_session_factory()
    if not factory:
        raise RuntimeError("DB가 초기화되지 않았습니다")
    async with factory() as session:
        rows = await session.execute(
            select(GpuDeviceCatalog).order_by(GpuDeviceCatalog.vendor_id, GpuDeviceCatalog.device_id)
        )
        return [_row_to_dict(r) for r in rows.scalars()]


async def upsert_device(vendor_id: str, device_id: str, name: str, is_audio: bool, aliases: list[str]) -> dict:
    """단건 추가/수정 (vendor_id+device_id 기준 upsert)."""
    vendor_id, device_id = validate_entry(vendor_id, device_id, name, aliases)
    factory = get_session_factory()
    if not factory:
        raise RuntimeError("DB가 초기화되지 않았습니다")
    async with factory() as session, session.begin():
        result = await session.execute(
            select(GpuDeviceCatalog).where(
                GpuDeviceCatalog.vendor_id == vendor_id, GpuDeviceCatalog.device_id == device_id
            )
        )
        row = result.scalar_one_or_none()
        if row:
            row.name = name
            row.is_audio = is_audio
            row.aliases = aliases
        else:
            row = GpuDeviceCatalog(
                vendor_id=vendor_id, device_id=device_id, name=name, is_audio=is_audio, aliases=aliases
            )
            session.add(row)
        await session.flush()
        return _row_to_dict(row)


async def delete_device(vendor_id: str, device_id: str) -> bool:
    """DB 항목 삭제. 존재하지 않으면 False."""
    factory = get_session_factory()
    if not factory:
        raise RuntimeError("DB가 초기화되지 않았습니다")
    async with factory() as session, session.begin():
        result = await session.execute(
            delete(GpuDeviceCatalog).where(
                GpuDeviceCatalog.vendor_id == vendor_id.upper(), GpuDeviceCatalog.device_id == device_id.upper()
            )
        )
        return result.rowcount > 0


async def bulk_import(entries: list[dict], mode: str = "replace") -> int:
    """카탈로그 일괄 import.

    mode="replace": DB 항목 전체 삭제 후 삽입 (단일 트랜잭션).
    mode="upsert": vendor_id+device_id 기준 추가/수정만, 기존 항목 유지.
    """
    if mode not in ("replace", "upsert"):
        raise ValueError(f"지원하지 않는 mode: {mode!r}")
    normalized = []
    for e in entries:
        vendor_id, device_id = validate_entry(e["vendor_id"], e["device_id"], e["name"], e["aliases"])
        normalized.append({**e, "vendor_id": vendor_id, "device_id": device_id})
    # 같은 vendor+device가 CSV 안에서 중복되면 마지막 행이 이긴다
    deduped = {(e["vendor_id"], e["device_id"]): e for e in normalized}

    factory = get_session_factory()
    if not factory:
        raise RuntimeError("DB가 초기화되지 않았습니다")
    async with factory() as session, session.begin():
        if mode == "replace":
            await session.execute(delete(GpuDeviceCatalog))
            for e in deduped.values():
                session.add(
                    GpuDeviceCatalog(
                        vendor_id=e["vendor_id"],
                        device_id=e["device_id"],
                        name=e["name"],
                        is_audio=bool(e["is_audio"]),
                        aliases=e["aliases"],
                    )
                )
        else:
            existing = {
                (r.vendor_id, r.device_id): r for r in (await session.execute(select(GpuDeviceCatalog))).scalars()
            }
            for key, e in deduped.items():
                row = existing.get(key)
                if row:
                    row.name = e["name"]
                    row.is_audio = bool(e["is_audio"])
                    row.aliases = e["aliases"]
                else:
                    session.add(
                        GpuDeviceCatalog(
                            vendor_id=e["vendor_id"],
                            device_id=e["device_id"],
                            name=e["name"],
                            is_audio=bool(e["is_audio"]),
                            aliases=e["aliases"],
                        )
                    )
    return len(deduped)


def _entry_from_cells(lineno: int, row: list[str]) -> dict | None:
    """행(셀 리스트)을 카탈로그 항목으로 변환. 빈 행/헤더 행이면 None.

    CSV와 xlsx 파싱이 같은 검증 경로를 타도록 공용으로 사용한다.
    """
    if not row or all(not c.strip() for c in row):
        return None
    if row[0].strip().lower() == "vendor_id":  # 헤더 행
        return None
    if len(row) < 4:
        raise ValueError(f"{lineno}행: 컬럼 수 부족 (vendor_id,device_id,name,is_audio,aliases 필요)")
    vendor_id = row[0].strip()
    device_id = row[1].strip()
    name = row[2].strip()
    is_audio_raw = row[3].strip().lower()
    if is_audio_raw not in ("true", "false", "1", "0", ""):
        raise ValueError(f"{lineno}행: is_audio는 true/false여야 합니다: {row[3]!r}")
    aliases_raw = row[4].strip() if len(row) > 4 else ""
    aliases = [a.strip() for a in aliases_raw.split(";") if a.strip()]
    try:
        validate_entry(vendor_id, device_id, name, aliases)
    except ValueError as e:
        raise ValueError(f"{lineno}행: {e}") from e
    return {
        "vendor_id": vendor_id,
        "device_id": device_id,
        "name": name,
        "is_audio": is_audio_raw in ("true", "1"),
        "aliases": aliases,
    }


def parse_csv(content: str) -> list[dict]:
    """CSV 텍스트를 카탈로그 항목 리스트로 파싱.

    포맷: vendor_id,device_id,name,is_audio,aliases
    aliases는 세미콜론(;) 구분. 헤더 행은 첫 컬럼이 "vendor_id"면 건너뛴다.
    오류 시 행 번호를 포함한 ValueError를 던진다.
    """
    entries: list[dict] = []
    reader = csv.reader(io.StringIO(content))
    for lineno, row in enumerate(reader, start=1):
        entry = _entry_from_cells(lineno, row)
        if entry:
            entries.append(entry)
    if not entries:
        raise ValueError("CSV에 유효한 항목이 없습니다")
    return entries


def parse_xlsx(data: bytes) -> list[dict]:
    """엑셀(xlsx) 파일을 카탈로그 항목 리스트로 파싱.

    첫 번째 시트의 컬럼 순서는 CSV와 동일: vendor_id, device_id, name, is_audio, aliases.
    export 파일의 source 같은 추가 컬럼은 무시한다.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"엑셀 파일을 읽을 수 없습니다: {e}") from e
    try:
        ws = wb.worksheets[0]
        entries: list[dict] = []
        for lineno, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = ["" if c is None else str(c) for c in row[:5]]
            entry = _entry_from_cells(lineno, cells)
            if entry:
                entries.append(entry)
    finally:
        wb.close()
    if not entries:
        raise ValueError("엑셀에 유효한 항목이 없습니다")
    return entries


EXPORT_COLUMNS = [*CSV_COLUMNS, "source"]


def build_export_csv(devices: list[dict]) -> str:
    """병합 카탈로그를 CSV 텍스트로 직렬화 (템플릿 다운로드용)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(EXPORT_COLUMNS)
    for d in devices:
        writer.writerow(
            [
                d["vendor_id"],
                d["device_id"],
                d["name"],
                "true" if d["is_audio"] else "false",
                ";".join(d.get("aliases", [])),
                d.get("source", ""),
            ]
        )
    return buf.getvalue()


def build_export_xlsx(devices: list[dict]) -> bytes:
    """병합 카탈로그를 엑셀(xlsx) 바이트로 직렬화 (템플릿 다운로드용)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "gpu_devices"
    ws.append(EXPORT_COLUMNS)
    for d in devices:
        ws.append(
            [
                d["vendor_id"],
                d["device_id"],
                d["name"],
                "true" if d["is_audio"] else "false",
                ";".join(d.get("aliases", [])),
                d.get("source", ""),
            ]
        )
    widths = {"A": 11, "B": 11, "C": 28, "D": 10, "E": 40, "F": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def refresh_device_map_from_db() -> None:
    """DB 카탈로그를 읽어 PCI_DEVICE_MAP에 overlay하고 관련 lazy 캐시를 무효화."""
    from app.services.gpu_inventory import apply_db_overlay
    from app.services.gpu_quota import invalidate_norm_map

    entries = await list_db_devices()
    apply_db_overlay(entries)
    invalidate_norm_map()
